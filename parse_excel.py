from openpyxl import load_workbook
import json

def fill_excel_data(template_path, row_name_with_data, excel_save_path):
    dataset = load_workbook(template_path)
    sheet = dataset.active
    sheet.protection.sheet = False
    merged_cells = sheet.merged_cells.ranges

    for key in row_name_with_data:
        data = row_name_with_data[key]
        if data is not None and 'value' in data:
            sheet.cell(data['row'], data['col'], data['value'])

    for merged_cell in merged_cells:
        sheet.merge_cells(str(merged_cell))
    dataset.save(excel_save_path)

def read_excel_data(file_path, excel_config):
    dataset = load_workbook(file_path)
    sheet = dataset.active
    sheet.protection.sheet = False
    mapping_data = {}
    for key in excel_config:
        excel_pos = excel_config[key]
        data = sheet.cell(excel_pos['row'], excel_pos['col']).internal_value
        if data is not None:
            mapping_data[key] = str(data).replace(",", "")
    return mapping_data


def parse_json_config(structure_file_path):
    with open(structure_file_path, 'r', encoding='utf-8') as file:
        try:
            return json.load(file)
        except json.JSONDecodeError as e:
            print(f"Error decoding JSON: {e}")
            return None
